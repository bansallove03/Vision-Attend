import os
import pickle
import numpy as np
import cv2
import face_recognition
import cvzone
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

# Load Encode File (containing encodings and IDs)
print("Loading Encode File ...")
file = open('EncodeFile.p', 'rb')
encodeListKnownWithIds = pickle.load(file)
file.close()
encodeListKnown, studentIds = encodeListKnownWithIds
print("Encode File Loaded")

# Excel file path
excel_file = 'StudentData.xlsx'

# Global variables for stopping the process
stop_attendance = False


def update_attendance(student_id):
    global temp

    # Load the existing data from Excel
    student_data = pd.read_excel(excel_file, engine='openpyxl')

    # Ensure no spaces and case consistency
    student_data['Roll Number'] = student_data['Roll Number'].str.strip().str.lower()
    student_id = student_id.lower()

    # Check if the student ID exists in the data
    if student_id in student_data['Roll Number'].values:
        student_info = student_data[student_data['Roll Number'] == student_id].iloc[0]
        print(f"Found student: {student_info['Name']} with Roll Number {student_id}")

        # Update Total Attendance
        student_data.loc[student_data['Roll Number'] == student_id, 'Total Attendance'] += 1
        student_data.loc[student_data['Roll Number'] == student_id, 'Last Attendance Time'] = pd.Timestamp.now()

        # Save the updated data back to Excel
        wb = load_workbook(excel_file)
        sheet = wb.active

        # Update the row corresponding to the student
        for idx, row in student_data.iterrows():
            if row['Roll Number'] == student_id:
                for col, value in enumerate(row, start=1):
                    sheet.cell(row=idx + 2, column=col, value=value)

         # Set the width of columns (can adjust these as per your requirement)
        sheet.column_dimensions['A'].width = 20  # Roll Number column
        sheet.column_dimensions['B'].width = 30  # Name column
        sheet.column_dimensions['C'].width = 20  # Total Attendance column
        sheet.column_dimensions['D'].width = 30  # Last Attendance Time column

        # Save the workbook with updated column widths
        wb.save(excel_file)
        temp = 0 #set the workbook with updated column widths
        return student_info
    else:
        temp = 1  # If student ID not found, set temp to 1
        print(f"Student ID {student_id} not found in the data.")
        return None


def run_face_attendance():
    """
    Main function to run the face attendance system.
    """
    global stop_attendance
    stop_attendance = False  # Reset the stop condition

    # Initialize webcama
    #cap = cv2.VideoCapture(0)
    #cap.set(3, 640)
    #cap.set(4, 480)

    cap = cv2.VideoCapture(1, cv2.CAP_DSHOW)
    cap.set(3, 640)
    cap.set(4, 480)

    # Load the background image
    imgBackground = cv2.imread('static/Resources/background.png')

    # Load the mode images
    folderModePath = 'static/Resources/Modes'
    modePathList = os.listdir(folderModePath)
    imgModeList = []
    for path in modePathList:
        imgModeList.append(cv2.imread(os.path.join(folderModePath, path)))

    counter = 0
    modeType = 0
    id = -1
    last_attendance = {}

    while True:
        # Check for stop condition
        if stop_attendance:
            print("Stopping attendance system...")
            cap.release()
            cv2.destroyAllWindows()
            break

        # Read video frame
        success, img = cap.read()
        if not success:
            print("Failed to capture video frame.")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        faceCurFrame = face_recognition.face_locations(imgS)
        encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        imgBackground[162:162 + 480, 55:55 + 640] = img  # Update background with video feed
        modeType = 0  # Default mode - no face detected

        if faceCurFrame:
            for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

                matchIndex = np.argmin(faceDis)

                if matches[matchIndex] and faceDis[matchIndex] < 0.45:  # Threshold
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
                    imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)
                    id = studentIds[matchIndex]


                    modeType = 2

                    if counter == 0:
                        cvzone.putTextRect(imgBackground, "Loading", (275, 400))
                        cv2.imshow("Face Attendance", imgBackground)
                        cv2.waitKey(1)
                        counter = 1
            if counter != 0:
                current_time = datetime.now()
                last_time = last_attendance.get(id, None)

                if last_time is None or (current_time - last_time) > timedelta(seconds=60): # minimum time 60 second
                    student_info = update_attendance(id)

                    if student_info is not None:
                        last_attendance[id] = current_time  # Update the last attendance time

                        if counter <= 10:
                            cv2.putText(imgBackground, str(student_info['Total Attendance']), (861, 125),
                                        cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                            cv2.putText(imgBackground, str(student_info['Name']), (808 + 50, 445),
                                        cv2.FONT_HERSHEY_COMPLEX, 1, (50, 50, 50), 1)

                            cvzone.putTextRect(imgBackground, "Marked",  (275, 450), colorR=(0, 255, 0))
                            cv2.putText(imgBackground, f"ID: {id}", (55 + x1, 162 + y1 - 10),  # Above the bounding box
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

                        counter += 1
                else:
                    modeType = 3  # Already marked mode
                    cvzone.putTextRect(imgBackground, "Already Marked", (275, 450), colorR=(0, 0, 255))
                    cv2.putText(imgBackground, f"ID: {id}", (55 + x1, 162 + y1 - 10),  # Above the bounding box
                                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

                if counter >= 20:
                    counter = 0
                    modeType = 0  # Reset to active mode
                    student_info = []
                    imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

        else:
            modeType = 0
            counter = 0

        imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

        # Resize the entire imgBackground after adding all text and overlays
        imgBackground_resized = cv2.resize(imgBackground,
                                           (int(imgBackground.shape[1] / 2), int(imgBackground.shape[0] / 2)))

        # Display the resized image
        cv2.imshow("Face Attendance", imgBackground_resized)
        if cv2.waitKey(1) == ord('q'):
            stop_attendance = True

    print("Attendance system stopped.")
